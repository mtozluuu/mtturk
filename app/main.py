import os
import warnings
from collections import defaultdict
from io import BytesIO
from contextlib import asynccontextmanager
from datetime import date, datetime

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, Form, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload
from starlette.middleware.sessions import SessionMiddleware
from starlette.status import HTTP_303_SEE_OTHER

from app.database import SessionLocal, get_db
from app.models import CrewAssignment, Flight, FlightNote, MaintenanceLog, User
from app.routers import admin, auth, flights, reports
from app.routers.auth import pwd_context
from app.seed import seed_admin

load_dotenv()

TRANSLATIONS = {
    "en": {
        "nav.home": "Home",
        "nav.flights": "Flights",
        "nav.new_flight": "New Flight",
        "nav.admin": "Admin Panel",
        "nav.reports": "Reports",
        "nav.api_docs": "API Docs",
        "nav.login": "Login",
        "nav.logout": "Logout",

        "msg.error.invalid_login": "Invalid username or password.",
        "msg.error.invalid_role": "Invalid role selected.",
        "msg.error.user_not_found": "User not found.",
        "msg.error.flight_not_found": "Flight not found.",
        "msg.error.invalid_seat": "Invalid seat selection.",
        "msg.error.role_mismatch": "Selected user roles are not compatible.",
        "msg.error.missing_crew_selection": "Select at least one crew member to change.",
        "msg.error.same_personnel": "Captain and first officer cannot be the same person.",
        "msg.error.inactive_flight": "Crew handover is only allowed for active flights.",
        "msg.error.empty_description": "Maintenance description cannot be empty.",
        "msg.error.username_exists": "This username already exists.",
        "msg.error.invalid_schedule_range": "Scheduled arrival cannot be earlier than scheduled departure.",
        "msg.error.invalid_actual_departure": "Actual departure cannot be earlier than scheduled departure.",
        "msg.error.invalid_actual_arrival": "Actual arrival cannot be earlier than scheduled arrival.",
        "msg.error.invalid_actual_range": "Actual arrival cannot be earlier than actual departure.",
        "msg.error.actual_departure_set": "Actual departure is already set for this flight.",
        "msg.error.actual_arrival_set": "Actual arrival is already set for this flight.",
        "msg.error.empty_flight_note": "Flight note cannot be empty.",
        "msg.error.invalid_crew_role": "Selected user must have pilot or copilot role.",
        "msg.error.arrival_before_departure": "Actual arrival cannot be earlier than actual departure.",

        "msg.success.login": "Logged in successfully.",
        "msg.success.logout": "Logged out successfully.",
        "msg.success.flight_created": "Flight created successfully.",
        "msg.success.flight_updated": "Flight updated successfully.",
        "msg.success.flight_deleted": "Flight deleted successfully.",
        "msg.success.user_created": "User created successfully.",
        "msg.success.role_updated": "User role updated successfully.",
        "msg.success.crew_updated": "Crew assignments updated successfully.",
        "msg.success.maintenance_created": "Maintenance record added successfully.",
        "msg.success.actual_departure_marked": "Actual departure time marked.",
        "msg.success.actual_arrival_marked": "Actual arrival time marked.",
        "msg.success.flight_note_created": "Flight note added successfully.",
        "nav.language": "Language",
        "lang.tr": "Türkçe",
        "lang.en": "English",
        "password.change": "Change Password",
        "password.current": "Current Password",
        "password.new": "New Password",
        "password.confirm": "Confirm New Password",
        "password.button": "Update Password",
        "password.error.wrong": "Current password is incorrect.",
        "password.error.mismatch": "New password and confirmation do not match.",
        "password.success": "Password updated successfully.",

        "reports.title": "Reports",
        "reports.description": "View system-wide flight, user, and active crew statistics.",
        "reports.crew_reports": "Crew Reports",
        "reports.general_stats": "General Statistics",
        "reports.total_flights": "Total Flights",
        "reports.today_flights": "Today's Flights",
        "reports.total_users": "Total Users",
        "reports.active_assignments": "Active Crew Assignments",
        "reports.roles": "User Roles",
        "reports.admins": "Admins",
        "reports.pilots": "Pilots",
        "reports.copilots": "Copilots",
        "reports.technicians": "Technicians",
        "reports.recent_flights": "Recent Flights",
        "reports.no_flights": "No flights have been recorded yet.",

        "crew_reports.title": "Crew Reports",
        "crew_reports.description": "View flight segments, total durations, and flight counts by pilot and copilot.",
        "crew_reports.summary_cards": "Summary Cards",
        "crew_reports.total_crew_minutes": "Total Crew Minutes",
        "crew_reports.total_crew_time": "Total Crew Time",
        "crew_reports.total_assignments": "Total Assignments",
        "crew_reports.total_unique_flights": "Total Unique Flights",
        "crew_reports.total_crew_users": "Total Crew Users",
        "crew_reports.filters": "Filters",
        "crew_reports.start_date": "Start Date",
        "crew_reports.end_date": "End Date",
        "crew_reports.personnel": "Personnel",
        "crew_reports.role": "Role",
        "crew_reports.all": "All",
        "crew_reports.filter": "Filter",
        "crew_reports.clear": "Clear",
        "crew_reports.export": "Excel Export",
        "crew_reports.summary": "Summary",
        "crew_reports.detail": "Detail",
        "crew_reports.no_records": "No records found.",
        "crew_reports.total_time_hms": "Total Time (HH:MM:SS)",
        "crew_reports.duration_hms": "Duration (HH:MM:SS)",
        "crew_reports.duration_min": "Duration (min)",
        "crew_reports.status": "Status",
        "crew_reports.active": "ACTIVE",
        "crew_reports.closed": "CLOSED",
        "crew_reports.route": "Route",
        "crew_reports.user": "User",
        "crew_reports.seat": "Seat",
        "crew_reports.start": "Start",
        "crew_reports.end": "End",
        "crew_reports.date": "Date",
        "crew_reports.flight_no": "Flight No",
        "crew_reports.total_minutes": "Total Minutes",
        "crew_reports.total_flights_col": "Total Flights",

        "crew_assignment.title": "Crew Assignment",
        "crew_assignment.flight": "Flight",
        "crew_assignment.date": "Date",
        "crew_assignment.route": "Route",
        "crew_assignment.scheduled": "Scheduled",
        "crew_assignment.actual": "Actual",
        "crew_assignment.active_crew": "Active Crew",
        "crew_assignment.no_active": "There is no active crew assignment for this flight.",
        "crew_assignment.bulk_assign": "Initial / Bulk Assignment",
        "crew_assignment.captain": "Captain",
        "crew_assignment.first_officer": "First Officer",
        "crew_assignment.bulk_assign_btn": "Assign All",
        "crew_assignment.handover": "Crew Change / Handover",
        "crew_assignment.new_personnel": "New Personnel",
        "crew_assignment.current_captain": "Current Captain",
        "crew_assignment.current_first_officer": "Current First Officer",
        "crew_assignment.handover_help_optional": "Captain and first officer selections are optional.",
        "crew_assignment.handover_help_empty": "Leave a field empty to keep the current assignment unchanged.",
        "crew_assignment.handover_help_required": "Select at least one role to apply a handover.",
        "crew_assignment.new_captain_optional": "New Captain (leave empty to keep current)",
        "crew_assignment.new_first_officer_optional": "New First Officer (leave empty to keep current)",
        "crew_assignment.keep_current": "Keep current assignment",
        "crew_assignment.save_change": "Apply Handover",
        "crew_assignment.back": "Back",
        "crew_assignment.history": "Assignment History",
        "crew_assignment.no_history": "There is no assignment history for this flight yet.",
        "crew_assignment.duration_min": "Duration (min)",
        "crew_assignment.duration_hms": "Duration (HH:MM:SS)",
        "crew_assignment.status": "Status",
        "crew_assignment.active": "ACTIVE",
        "crew_assignment.closed": "CLOSED",

        "flight_detail.title": "Flight Operations",
        "flight_detail.flight": "Flight",
        "flight_detail.date": "Date",
        "flight_detail.route": "Route",
        "flight_detail.planned_departure": "Planned Departure",
        "flight_detail.planned_arrival": "Planned Arrival",
        "flight_detail.actual_departure": "Actual Departure",
        "flight_detail.actual_arrival": "Actual Arrival",
        "flight_detail.not_set": "Not set",
        "flight_detail.mark_departure_now": "Mark Departure Now",
        "flight_detail.mark_arrival_now": "Mark Arrival Now",
        "flight_detail.crew": "Active Crew",
        "flight_detail.current_captain": "Current Captain",
        "flight_detail.current_first_officer": "Current First Officer",
        "flight_detail.assignments": "Crew Assignment",
        "flight_detail.captain": "Captain",
        "flight_detail.first_officer": "First Officer",
        "flight_detail.select_user": "Select user",
        "flight_detail.save_assignments": "Update Crew",
        "flight_detail.notes": "Flight Notes",
        "flight_detail.new_note": "New Note",
        "flight_detail.note_placeholder": "Write an operational note...",
        "flight_detail.add_note": "Add Note",
        "flight_detail.no_notes": "No flight notes yet.",
        "flight_detail.unknown_user": "User",
    },
    "tr": {
        "nav.home": "Ana Sayfa",
        "nav.flights": "Uçuşlar",
        "nav.new_flight": "Yeni Uçuş Ekle",
        "nav.admin": "Admin Panel",
        "nav.reports": "Raporlar",
        "nav.api_docs": "API Dokümantasyonu",
        "nav.login": "Giriş",
        "nav.logout": "Çıkış",

        "msg.error.invalid_login": "Kullanıcı adı veya şifre hatalı.",
        "msg.error.invalid_role": "Geçersiz rol seçildi.",
        "msg.error.user_not_found": "Kullanıcı bulunamadı.",
        "msg.error.flight_not_found": "Uçuş bulunamadı.",
        "msg.error.invalid_seat": "Geçersiz koltuk seçimi.",
        "msg.error.role_mismatch": "Seçilen kullanıcı rolleri uygun değil.",
        "msg.error.missing_crew_selection": "Değişiklik için en az bir ekip üyesi seçin.",
        "msg.error.same_personnel": "Kaptan ve yardımcı pilot aynı kişi olamaz.",
        "msg.error.inactive_flight": "Ekip devri yalnızca aktif uçuşlarda yapılabilir.",
        "msg.error.empty_description": "Bakım açıklaması boş bırakılamaz.",
        "msg.error.username_exists": "Bu kullanıcı adı zaten mevcut.",
        "msg.error.invalid_schedule_range": "Planlanan varış, planlanan kalkıştan önce olamaz.",
        "msg.error.invalid_actual_departure": "Gerçek kalkış, planlanan kalkıştan önce olamaz.",
        "msg.error.invalid_actual_arrival": "Gerçek varış, planlanan varıştan önce olamaz.",
        "msg.error.invalid_actual_range": "Gerçek varış, gerçek kalkıştan önce olamaz.",
        "msg.error.actual_departure_set": "Bu uçuş için gerçek kalkış zaten işaretlenmiş.",
        "msg.error.actual_arrival_set": "Bu uçuş için gerçek varış zaten işaretlenmiş.",
        "msg.error.empty_flight_note": "Uçuş notu boş bırakılamaz.",
        "msg.error.invalid_crew_role": "Seçilen kullanıcının rolü pilot veya yardımcı pilot olmalıdır.",
        "msg.error.arrival_before_departure": "Gerçek varış, gerçek kalkıştan önce olamaz.",

        "msg.success.login": "Başarıyla giriş yapıldı.",
        "msg.success.logout": "Başarıyla çıkış yapıldı.",
        "msg.success.flight_created": "Uçuş başarıyla oluşturuldu.",
        "msg.success.flight_updated": "Uçuş başarıyla güncellendi.",
        "msg.success.flight_deleted": "Uçuş başarıyla silindi.",
        "msg.success.user_created": "Kullanıcı başarıyla oluşturuldu.",
        "msg.success.role_updated": "Kullanıcı rolü başarıyla güncellendi.",
        "msg.success.crew_updated": "Ekip atamaları başarıyla güncellendi.",
        "msg.success.maintenance_created": "Bakım kaydı başarıyla eklendi.",
        "msg.success.actual_departure_marked": "Gerçek kalkış saati işaretlendi.",
        "msg.success.actual_arrival_marked": "Gerçek varış saati işaretlendi.",
        "msg.success.flight_note_created": "Uçuş notu başarıyla eklendi.",

        "nav.language": "Dil",
        "lang.tr": "Türkçe",
        "lang.en": "English",
        "password.change": "Şifre Değiştir",
        "password.current": "Mevcut Şifre",
        "password.new": "Yeni Şifre",
        "password.confirm": "Yeni Şifre Tekrar",
        "password.button": "Şifreyi Güncelle",
        "password.error.wrong": "Mevcut şifre yanlış.",
        "password.error.mismatch": "Yeni şifre ve tekrar alanı eşleşmiyor.",
        "password.success": "Şifre başarıyla güncellendi.",

        "reports.title": "Raporlar",
        "reports.description": "Sistem genelindeki uçuş, kullanıcı ve aktif ekip istatistiklerini görüntüleyin.",
        "reports.crew_reports": "Ekip Raporları",
        "reports.general_stats": "Genel İstatistikler",
        "reports.total_flights": "Toplam Uçuş",
        "reports.today_flights": "Bugünkü Uçuşlar",
        "reports.total_users": "Toplam Kullanıcı",
        "reports.active_assignments": "Aktif Ekip Atamaları",
        "reports.roles": "Kullanıcı Rolleri",
        "reports.admins": "Yöneticiler",
        "reports.pilots": "Pilotlar",
        "reports.copilots": "Yardımcı Pilotlar",
        "reports.technicians": "Teknisyenler",
        "reports.recent_flights": "Son Uçuşlar",
        "reports.no_flights": "Henüz kayıtlı uçuş bulunmuyor.",

        "crew_reports.title": "Ekip Raporları",
        "crew_reports.description": "Pilot ve yardımcı pilot bazlı uçuş segmentlerini, toplam süreleri ve uçuş sayılarını görüntüleyin.",
        "crew_reports.summary_cards": "Özet Kartlar",
        "crew_reports.total_crew_minutes": "Toplam Ekip Dakikası",
        "crew_reports.total_crew_time": "Toplam Ekip Süresi",
        "crew_reports.total_assignments": "Toplam Atama",
        "crew_reports.total_unique_flights": "Toplam Benzersiz Uçuş",
        "crew_reports.total_crew_users": "Toplam Ekip Kullanıcısı",
        "crew_reports.filters": "Filtreler",
        "crew_reports.start_date": "Başlangıç Tarihi",
        "crew_reports.end_date": "Bitiş Tarihi",
        "crew_reports.personnel": "Personel",
        "crew_reports.role": "Rol",
        "crew_reports.all": "Tümü",
        "crew_reports.filter": "Filtrele",
        "crew_reports.clear": "Temizle",
        "crew_reports.export": "Excel Aktar",
        "crew_reports.summary": "Özet",
        "crew_reports.detail": "Detay",
        "crew_reports.no_records": "Kayıt bulunamadı.",
        "crew_reports.total_time_hms": "Toplam Süre (SS:DD:SN)",
        "crew_reports.duration_hms": "Süre (SS:DD:SN)",
        "crew_reports.duration_min": "Süre (dk)",
        "crew_reports.status": "Durum",
        "crew_reports.active": "AKTİF",
        "crew_reports.closed": "KAPALI",
        "crew_reports.route": "Rota",
        "crew_reports.user": "Kullanıcı",
        "crew_reports.seat": "Koltuk",
        "crew_reports.start": "Başlangıç",
        "crew_reports.end": "Bitiş",
        "crew_reports.date": "Tarih",
        "crew_reports.flight_no": "Uçuş No",
        "crew_reports.total_minutes": "Toplam Dakika",
        "crew_reports.total_flights_col": "Toplam Uçuş",

        "crew_assignment.title": "Ekip Ataması",
        "crew_assignment.flight": "Uçuş",
        "crew_assignment.date": "Tarih",
        "crew_assignment.route": "Rota",
        "crew_assignment.scheduled": "Planlanan",
        "crew_assignment.actual": "Gerçekleşen",
        "crew_assignment.active_crew": "Aktif Ekip",
        "crew_assignment.no_active": "Bu uçuş için aktif ekip ataması yok.",
        "crew_assignment.bulk_assign": "İlk / Toplu Atama",
        "crew_assignment.captain": "Kaptan",
        "crew_assignment.first_officer": "Yardımcı Pilot",
        "crew_assignment.bulk_assign_btn": "Toplu Ata",
        "crew_assignment.handover": "Ekip Değişimi / Devir",
        "crew_assignment.new_personnel": "Yeni Personel",
        "crew_assignment.current_captain": "Mevcut Kaptan",
        "crew_assignment.current_first_officer": "Mevcut Yardımcı Pilot",
        "crew_assignment.handover_help_optional": "Kaptan ve yardımcı pilot seçimleri isteğe bağlıdır.",
        "crew_assignment.handover_help_empty": "Boş bırakılan alan, mevcut atamanın korunacağı anlamına gelir.",
        "crew_assignment.handover_help_required": "Devir uygulamak için en az bir rol seçin.",
        "crew_assignment.new_captain_optional": "Yeni Kaptan (aynı kalacaksa boş bırakın)",
        "crew_assignment.new_first_officer_optional": "Yeni Yardımcı Pilot (aynı kalacaksa boş bırakın)",
        "crew_assignment.keep_current": "Mevcut atamayı koru",
        "crew_assignment.save_change": "Ekip Devrini Uygula",
        "crew_assignment.back": "Geri Dön",
        "crew_assignment.history": "Atama Geçmişi",
        "crew_assignment.no_history": "Bu uçuş için henüz atama geçmişi bulunmuyor.",
        "crew_assignment.duration_min": "Süre (dk)",
        "crew_assignment.duration_hms": "Süre (SS:DD:SN)",
        "crew_assignment.status": "Durum",
        "crew_assignment.active": "AKTİF",
        "crew_assignment.closed": "KAPALI",

        "flight_detail.title": "Uçuş Operasyonu",
        "flight_detail.flight": "Uçuş",
        "flight_detail.date": "Tarih",
        "flight_detail.route": "Rota",
        "flight_detail.planned_departure": "Planlanan Kalkış",
        "flight_detail.planned_arrival": "Planlanan Varış",
        "flight_detail.actual_departure": "Gerçek Kalkış",
        "flight_detail.actual_arrival": "Gerçek Varış",
        "flight_detail.not_set": "İşaretlenmedi",
        "flight_detail.mark_departure_now": "Kalkışı Şimdi İşaretle",
        "flight_detail.mark_arrival_now": "Varışı Şimdi İşaretle",
        "flight_detail.crew": "Aktif Ekip",
        "flight_detail.current_captain": "Mevcut Kaptan",
        "flight_detail.current_first_officer": "Mevcut Yardımcı Pilot",
        "flight_detail.assignments": "Ekip Ataması",
        "flight_detail.captain": "Kaptan",
        "flight_detail.first_officer": "Yardımcı Pilot",
        "flight_detail.select_user": "Kullanıcı seçin",
        "flight_detail.save_assignments": "Ekibi Güncelle",
        "flight_detail.notes": "Uçuş Notları",
        "flight_detail.new_note": "Yeni Not",
        "flight_detail.note_placeholder": "Operasyon notu yazın...",
        "flight_detail.add_note": "Not Ekle",
        "flight_detail.no_notes": "Henüz uçuş notu yok.",
        "flight_detail.unknown_user": "Kullanıcı",
    },
}

SESSION_SECRET_KEY = os.environ.get("SESSION_SECRET_KEY", "")
if not SESSION_SECRET_KEY:
    warnings.warn(
        "SESSION_SECRET_KEY is not set. Using an insecure default. "
        "Set SESSION_SECRET_KEY in your environment before deploying.",
        stacklevel=1,
    )
    SESSION_SECRET_KEY = "insecure-dev-key-do-not-use-in-production"

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))


@asynccontextmanager
async def lifespan(app: FastAPI):
    seed_admin()
    yield


def get_session_user(request: Request) -> User | None:
    if "session" not in request.scope:
        return None

    user_id = request.session.get("user_id")
    if not user_id:
        return None

    db = SessionLocal()
    try:
        return db.query(User).filter(User.id == user_id).first()
    finally:
        db.close()


def get_locale(request: Request) -> str:
    if "session" not in request.scope:
        return "tr"

    lang = request.session.get("lang", "tr")
    if lang not in TRANSLATIONS:
        return "tr"
    return lang

def i18n_ctx(request, extra=None):
    lang = get_locale(request)
    t = lambda k: TRANSLATIONS[lang].get(k, k)
    ctx = {"t": t, "lang": lang, "request": request}
    if extra:
        ctx.update(extra)
    return ctx


def translate(request: Request, key: str) -> str:
    lang = get_locale(request)
    return TRANSLATIONS.get(lang, {}).get(key, key)


def parse_optional_user_id_form_value(raw_value: str) -> int | None:
    value = (raw_value or "").strip()
    if not value:
        return None
    if not value.isdigit():
        raise ValueError
    return int(value)


def resolve_assignment_end_time(assignment: CrewAssignment, flight: Flight) -> datetime:
    if assignment.end_time is not None:
        return assignment.end_time
    if flight.actual_arr is not None:
        return flight.actual_arr
    return datetime.utcnow()


def assignment_duration_seconds(assignment: CrewAssignment, flight: Flight) -> int:
    end_time = resolve_assignment_end_time(assignment, flight)
    delta = end_time - assignment.start_time
    return max(int(delta.total_seconds()), 0)


def assignment_duration_minutes(assignment: CrewAssignment, flight: Flight) -> int:
    return assignment_duration_seconds(assignment, flight) // 60


def format_duration_hms(total_seconds: int) -> str:
    total_seconds = max(int(total_seconds), 0)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02}:{minutes:02}:{seconds:02}"


def build_crew_report_data(
    db: Session,
    start_date_str: str | None = None,
    end_date_str: str | None = None,
    user_id: int | None = None,
    role: str | None = None,
):
    query = (
        db.query(CrewAssignment)
        .join(Flight, Flight.id == CrewAssignment.flight_id)
        .join(User, User.id == CrewAssignment.user_id)
        .options(
            joinedload(CrewAssignment.user),
            joinedload(CrewAssignment.flight),
        )
        .order_by(Flight.flight_date.desc(), CrewAssignment.start_time.desc())
    )

    if start_date_str:
        start_date = date.fromisoformat(start_date_str)
        query = query.filter(Flight.flight_date >= start_date)

    if end_date_str:
        end_date = date.fromisoformat(end_date_str)
        query = query.filter(Flight.flight_date <= end_date)

    if user_id:
        query = query.filter(CrewAssignment.user_id == user_id)

    if role:
        query = query.filter(User.role == role)

    assignments = query.all()

    detail_rows = []
    summary_map = defaultdict(
        lambda: {
            "user_id": None,
            "username": "",
            "role": "",
            "flight_ids": set(),
            "total_assignments": 0,
            "total_seconds": 0,
        }
    )

    for assignment in assignments:
        flight = assignment.flight
        crew_user = assignment.user

        if not flight or not crew_user:
            continue

        duration_seconds = assignment_duration_seconds(assignment, flight)
        duration_minutes = duration_seconds // 60
        duration_hms = format_duration_hms(duration_seconds)

        detail_rows.append(
            {
                "flight_id": flight.id,
                "flight_no": flight.flight_no,
                "flight_date": flight.flight_date,
                "departure_airport": flight.departure_airport,
                "arrival_airport": flight.arrival_airport,
                "user_id": crew_user.id,
                "username": crew_user.username,
                "role": crew_user.role,
                "seat": assignment.seat,
                "start_time": assignment.start_time,
                "end_time": assignment.end_time,
                "duration_seconds": duration_seconds,
                "duration_minutes": duration_minutes,
                "duration_hms": duration_hms,
                "is_active": assignment.end_time is None,
            }
        )

        key = crew_user.id
        summary_map[key]["user_id"] = crew_user.id
        summary_map[key]["username"] = crew_user.username
        summary_map[key]["role"] = crew_user.role
        summary_map[key]["flight_ids"].add(flight.id)
        summary_map[key]["total_assignments"] += 1
        summary_map[key]["total_seconds"] += duration_seconds

    summary_rows = []
    for item in summary_map.values():
        summary_rows.append(
            {
                "user_id": item["user_id"],
                "username": item["username"],
                "role": item["role"],
                "total_flights": len(item["flight_ids"]),
                "total_assignments": item["total_assignments"],
                "total_seconds": item["total_seconds"],
                "total_minutes": item["total_seconds"] // 60,
                "total_hms": format_duration_hms(item["total_seconds"]),
            }
        )

    summary_rows.sort(key=lambda x: (-x["total_seconds"], x["username"].lower()))

    return summary_rows, detail_rows


app = FastAPI(title="Flight Management API", version="1.0.0", lifespan=lifespan)

app.mount("/static", StaticFiles(directory=os.path.join(_BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(_BASE_DIR, "templates"))


templates.env.globals["get_locale"] = get_locale

app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET_KEY,
    session_cookie="session",
    same_site="lax",
    https_only=False,
)

# Routers (JSON API)
app.include_router(auth.router)
app.include_router(admin.router)
app.include_router(flights.router)
app.include_router(reports.router)


@app.get("/", include_in_schema=False)
def index(request: Request):
    user = get_session_user(request)
    return templates.TemplateResponse(request, "index.html", i18n_ctx(request, {"user": user}))

@app.get("/login", include_in_schema=False)
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", i18n_ctx(request))

@app.post("/set-language", include_in_schema=False)
def set_language(
    request: Request,
    lang: str = Form(...),
    next_url: str = Form("/"),
):
    if lang not in TRANSLATIONS:
        lang = "tr"

    if "session" in request.scope:
        request.session["lang"] = lang

    if not next_url.startswith("/"):
        next_url = "/"

    return RedirectResponse(url=next_url, status_code=HTTP_303_SEE_OTHER)

@app.post("/login", include_in_schema=False)
def login_form(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.username == username).first()
    if user is None or not pwd_context.verify(password, user.password_hash):
        return RedirectResponse(url="/login?error=1", status_code=HTTP_303_SEE_OTHER)

    request.session["user_id"] = user.id
    return RedirectResponse(url="/flights-ui?success=login", status_code=HTTP_303_SEE_OTHER)

@app.post("/logout", include_in_schema=False)
def logout(request: Request):
    if "session" in request.scope:
        request.session.clear()
    return RedirectResponse(url="/?success=logout", status_code=HTTP_303_SEE_OTHER)

@app.get("/flights-ui", include_in_schema=False)
def flights_ui(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    _db = SessionLocal()
    try:
        flight_list = _db.query(Flight).order_by(Flight.sched_dep).all()
    finally:
        _db.close()

    return templates.TemplateResponse(
        request,
        "flights.html",
        i18n_ctx(request, {"flights": flight_list, "user": user}),
    )

@app.get("/flights-ui/new", include_in_schema=False)
def create_flight_page(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    return templates.TemplateResponse(request, "create_flight.html", i18n_ctx(request, {"user": user}))

@app.post("/flights-ui/new", include_in_schema=False)
def create_flight_from_form(
    request: Request,
    flight_no: str = Form(...),
    flight_date: str = Form(...),
    departure_airport: str = Form(...),
    arrival_airport: str = Form(...),
    sched_dep: str = Form(...),
    sched_arr: str = Form(...),
    actual_dep: str = Form(""),
    actual_arr: str = Form(""),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    flight = Flight(
        flight_no=flight_no.strip(),
        flight_date=date.fromisoformat(flight_date),
        departure_airport=departure_airport.strip().upper(),
        arrival_airport=arrival_airport.strip().upper(),
        sched_dep=datetime.fromisoformat(sched_dep),
        sched_arr=datetime.fromisoformat(sched_arr),
        actual_dep=datetime.fromisoformat(actual_dep) if actual_dep else None,
        actual_arr=datetime.fromisoformat(actual_arr) if actual_arr else None,
    )
    db.add(flight)
    db.commit()
    db.refresh(flight)

    return RedirectResponse(url="/flights-ui?success=flight_created", status_code=HTTP_303_SEE_OTHER)

@app.get("/flight-detail/{flight_id}", include_in_schema=False)
def flight_detail(request: Request, flight_id: int):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        flight = db.query(Flight).filter(Flight.id == flight_id).first()
        if not flight:
            return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)
        captain_assignment = (
            db.query(CrewAssignment)
            .options(joinedload(CrewAssignment.user))
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.seat == "CAPTAIN",
                CrewAssignment.end_time.is_(None),
            )
            .order_by(CrewAssignment.start_time.desc())
            .first()
        )
        first_officer_assignment = (
            db.query(CrewAssignment)
            .options(joinedload(CrewAssignment.user))
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.seat == "FIRST_OFFICER",
                CrewAssignment.end_time.is_(None),
            )
            .order_by(CrewAssignment.start_time.desc())
            .first()
        )
        crew_candidates = (
            db.query(User)
            .filter(User.role.in_(["pilot", "copilot"]))
            .order_by(User.username)
            .all()
        )
        flight_notes = (
            db.query(FlightNote)
            .options(joinedload(FlightNote.user))
            .filter(FlightNote.flight_id == flight_id)
            .order_by(FlightNote.created_at.desc())
            .all()
        )
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "flight_detail.html",
        i18n_ctx(request, {
            "user": user,
            "flight": flight,
            "captain_assignment": captain_assignment,
            "first_officer_assignment": first_officer_assignment,
            "crew_candidates": crew_candidates,
            "flight_notes": flight_notes,
        }),
    )


@app.post("/flight-detail/{flight_id}/mark-departure", include_in_schema=False)
def mark_flight_departure_now(
    flight_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    if flight.actual_dep is not None:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=actual_departure_set",
            status_code=HTTP_303_SEE_OTHER,
        )

    now = datetime.utcnow()
    if flight.actual_arr is not None and now > flight.actual_arr:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=arrival_before_departure",
            status_code=HTTP_303_SEE_OTHER,
        )

    flight.actual_dep = now
    db.commit()

    return RedirectResponse(
        url=f"/flight-detail/{flight_id}?success=actual_departure_marked",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.post("/flight-detail/{flight_id}/mark-arrival", include_in_schema=False)
def mark_flight_arrival_now(
    flight_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    if flight.actual_arr is not None:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=actual_arrival_set",
            status_code=HTTP_303_SEE_OTHER,
        )

    now = datetime.utcnow()
    if flight.actual_dep is not None and now < flight.actual_dep:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=arrival_before_departure",
            status_code=HTTP_303_SEE_OTHER,
        )

    flight.actual_arr = now
    open_assignments = (
        db.query(CrewAssignment)
        .filter(
            CrewAssignment.flight_id == flight_id,
            CrewAssignment.end_time.is_(None),
        )
        .all()
    )
    for assignment in open_assignments:
        if now >= assignment.start_time:
            assignment.end_time = now

    db.commit()

    return RedirectResponse(
        url=f"/flight-detail/{flight_id}?success=actual_arrival_marked",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.post("/flight-detail/{flight_id}/notes", include_in_schema=False)
def add_flight_note(
    flight_id: int,
    request: Request,
    note: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    stripped_note = note.strip()
    if not stripped_note:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=empty_flight_note",
            status_code=HTTP_303_SEE_OTHER,
        )

    db.add(
        FlightNote(
            flight_id=flight_id,
            user_id=user.id,
            note=stripped_note,
        )
    )
    db.commit()

    return RedirectResponse(
        url=f"/flight-detail/{flight_id}?success=flight_note_created",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.post("/flight-detail/{flight_id}/crew", include_in_schema=False)
def update_flight_detail_crew(
    flight_id: int,
    request: Request,
    captain_user_id: str = Form(...),
    first_officer_user_id: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    try:
        parsed_captain_user_id = parse_optional_user_id_form_value(captain_user_id)
        parsed_first_officer_user_id = parse_optional_user_id_form_value(first_officer_user_id)
    except ValueError:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=user_not_found",
            status_code=HTTP_303_SEE_OTHER,
        )

    if parsed_captain_user_id is None or parsed_first_officer_user_id is None:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=missing_crew_selection",
            status_code=HTTP_303_SEE_OTHER,
        )

    if parsed_captain_user_id == parsed_first_officer_user_id:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=same_personnel",
            status_code=HTTP_303_SEE_OTHER,
        )

    selected_users = (
        db.query(User)
        .filter(User.id.in_([parsed_captain_user_id, parsed_first_officer_user_id]))
        .all()
    )
    selected_user_map = {selected_user.id: selected_user for selected_user in selected_users}
    captain_user = selected_user_map.get(parsed_captain_user_id)
    first_officer_user = selected_user_map.get(parsed_first_officer_user_id)

    if captain_user is None or first_officer_user is None:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=user_not_found",
            status_code=HTTP_303_SEE_OTHER,
        )

    allowed_roles = {"pilot", "copilot"}
    if captain_user.role not in allowed_roles or first_officer_user.role not in allowed_roles:
        return RedirectResponse(
            url=f"/flight-detail/{flight_id}?error=invalid_crew_role",
            status_code=HTTP_303_SEE_OTHER,
        )

    now = datetime.utcnow()

    active_captain = (
        db.query(CrewAssignment)
        .filter(
            CrewAssignment.flight_id == flight_id,
            CrewAssignment.seat == "CAPTAIN",
            CrewAssignment.end_time.is_(None),
        )
        .first()
    )
    active_first_officer = (
        db.query(CrewAssignment)
        .filter(
            CrewAssignment.flight_id == flight_id,
            CrewAssignment.seat == "FIRST_OFFICER",
            CrewAssignment.end_time.is_(None),
        )
        .first()
    )

    if active_captain is None or active_captain.user_id != parsed_captain_user_id:
        if active_captain is not None:
            active_captain.end_time = now
        db.add(
            CrewAssignment(
                flight_id=flight_id,
                user_id=parsed_captain_user_id,
                seat="CAPTAIN",
                start_time=now,
            )
        )

    if active_first_officer is None or active_first_officer.user_id != parsed_first_officer_user_id:
        if active_first_officer is not None:
            active_first_officer.end_time = now
        db.add(
            CrewAssignment(
                flight_id=flight_id,
                user_id=parsed_first_officer_user_id,
                seat="FIRST_OFFICER",
                start_time=now,
            )
        )

    db.commit()
    return RedirectResponse(
        url=f"/flight-detail/{flight_id}?success=crew_updated",
        status_code=HTTP_303_SEE_OTHER,
    )

@app.get("/admin-ui", include_in_schema=False)
def admin_ui(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        users = db.query(User).order_by(User.id).all()
        flights = db.query(Flight).order_by(Flight.id.desc()).all()
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "admin.html",
        i18n_ctx(request, {
            "user": user,
            "users": users,
            "flights": flights,
        }),
    )

@app.get("/admin-ui/users/new", include_in_schema=False)
def create_user_page(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    return templates.TemplateResponse(
        request,
        "create_user.html",
        i18n_ctx(request, {"user": user}),
    )

@app.post("/admin-ui/users/new", include_in_schema=False)
def create_user_from_form(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    username = username.strip()
    role = role.strip().lower()

    valid_roles = {"admin", "pilot", "copilot", "technician"}
    if role not in valid_roles:
        return RedirectResponse(url="/admin-ui/users/new?error=invalid_role", status_code=HTTP_303_SEE_OTHER)

    existing = db.query(User).filter(User.username == username).first()
    if existing is not None:
        return RedirectResponse(url="/admin-ui/users/new?error=username_exists", status_code=HTTP_303_SEE_OTHER)

    new_user = User(
        username=username,
        password_hash=pwd_context.hash(password),
        role=role,
    )
    db.add(new_user)
    db.commit()

    return RedirectResponse(url="/admin-ui?success=user_created", status_code=HTTP_303_SEE_OTHER)



@app.post("/admin-ui/users/{user_id}/role", include_in_schema=False)
def update_user_role(
    user_id: int,
    request: Request,
    role: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if current_user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    role = role.strip().lower()
    valid_roles = {"admin", "pilot", "copilot", "technician"}
    if role not in valid_roles:
        return RedirectResponse(url="/admin-ui?error=invalid_role", status_code=HTTP_303_SEE_OTHER)

    target_user = db.query(User).filter(User.id == user_id).first()
    if target_user is None:
        return RedirectResponse(url="/admin-ui?error=user_not_found", status_code=HTTP_303_SEE_OTHER)

    target_user.role = role
    db.commit()

    return RedirectResponse(url="/admin-ui?success=role_updated", status_code=HTTP_303_SEE_OTHER)


@app.get("/admin-ui/flights/{flight_id}/crew", include_in_schema=False)
def manage_flight_crew_page(flight_id: int, request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        flight = db.query(Flight).filter(Flight.id == flight_id).first()
        if flight is None:
            return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

        crew_users = (
            db.query(User)
            .filter(User.role.in_(["pilot", "copilot"]))
            .order_by(User.username)
            .all()
        )

        active_assignments = (
            db.query(CrewAssignment)
            .options(joinedload(CrewAssignment.user))
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.end_time.is_(None),
            )
            .order_by(CrewAssignment.seat)
            .all()
        )

        assignment_history = (
            db.query(CrewAssignment)
            .options(joinedload(CrewAssignment.user))
            .filter(CrewAssignment.flight_id == flight_id)
            .order_by(CrewAssignment.start_time.desc())
            .all()
        )

        history_rows = []
        for assignment in assignment_history:
            duration_seconds = assignment_duration_seconds(assignment, flight)
            duration_minutes = duration_seconds // 60

            history_rows.append(
                {
                    "id": assignment.id,
                    "username": assignment.user.username if assignment.user else f"User #{assignment.user_id}",
                    "role": assignment.user.role if assignment.user else "-",
                    "user_id": assignment.user_id,
                    "seat": assignment.seat,
                    "start_time": assignment.start_time,
                    "end_time": assignment.end_time,
                    "duration_minutes": duration_minutes,
                    "duration_hms": format_duration_hms(duration_seconds),
                    "is_active": assignment.end_time is None,
                }
            )
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "crew_assignment.html",
        i18n_ctx(
            request,
            {
                "user": user,
                "flight": flight,
                "pilots": crew_users,
                "copilots": crew_users,
                "active_assignments": active_assignments,
                "assignment_history": history_rows,
            },
        ),
    )


@app.post("/admin-ui/flights/{flight_id}/crew", include_in_schema=False)
def assign_flight_crew(
    flight_id: int,
    request: Request,
    captain_user_id: str = Form(""),
    first_officer_user_id: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if current_user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    try:
        parsed_captain_user_id = parse_optional_user_id_form_value(captain_user_id)
        parsed_first_officer_user_id = parse_optional_user_id_form_value(first_officer_user_id)
    except ValueError:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=user_not_found",
            status_code=HTTP_303_SEE_OTHER,
        )

    if parsed_captain_user_id is None and parsed_first_officer_user_id is None:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=missing_crew_selection",
            status_code=HTTP_303_SEE_OTHER,
        )

    if (
        parsed_captain_user_id is not None
        and parsed_first_officer_user_id is not None
        and parsed_captain_user_id == parsed_first_officer_user_id
    ):
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=same_personnel",
            status_code=HTTP_303_SEE_OTHER,
        )

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    if flight.actual_dep is None or flight.actual_arr is not None:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=inactive_flight",
            status_code=HTTP_303_SEE_OTHER,
        )

    captain_user = None
    if parsed_captain_user_id is not None:
        captain_user = db.query(User).filter(User.id == parsed_captain_user_id).first()

    first_officer_user = None
    if parsed_first_officer_user_id is not None:
        first_officer_user = db.query(User).filter(User.id == parsed_first_officer_user_id).first()

    if (
        (parsed_captain_user_id is not None and captain_user is None)
        or (parsed_first_officer_user_id is not None and first_officer_user is None)
    ):
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=user_not_found",
            status_code=HTTP_303_SEE_OTHER,
        )

    if (
        (captain_user is not None and captain_user.role not in {"pilot", "copilot"})
        or (first_officer_user is not None and first_officer_user.role not in {"pilot", "copilot"})
    ):
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=invalid_crew_role",
            status_code=HTTP_303_SEE_OTHER,
        )

    now = datetime.utcnow()
    handover_performed = False

    if parsed_captain_user_id is not None:
        active_captain = (
            db.query(CrewAssignment)
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.seat == "CAPTAIN",
                CrewAssignment.end_time.is_(None),
            )
            .first()
        )
        if active_captain is None or active_captain.user_id != parsed_captain_user_id:
            if active_captain:
                active_captain.end_time = now
            db.add(
                CrewAssignment(
                    flight_id=flight_id,
                    user_id=parsed_captain_user_id,
                    seat="CAPTAIN",
                    start_time=now,
                )
            )
            handover_performed = True

    if parsed_first_officer_user_id is not None:
        active_first_officer = (
            db.query(CrewAssignment)
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.seat == "FIRST_OFFICER",
                CrewAssignment.end_time.is_(None),
            )
            .first()
        )
        if active_first_officer is None or active_first_officer.user_id != parsed_first_officer_user_id:
            if active_first_officer:
                active_first_officer.end_time = now
            db.add(
                CrewAssignment(
                    flight_id=flight_id,
                    user_id=parsed_first_officer_user_id,
                    seat="FIRST_OFFICER",
                    start_time=now,
                )
            )
            handover_performed = True

    db.commit()

    return RedirectResponse(
        url=f"/admin-ui/flights/{flight_id}/crew?success=crew_updated",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.post("/admin-ui/flights/{flight_id}/crew/change", include_in_schema=False)
def change_flight_crew(
    flight_id: int,
    request: Request,
    seat: str = Form(...),
    new_user_id: int = Form(...),
    db: Session = Depends(get_db),
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if current_user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    seat = seat.strip().upper()
    if seat not in {"CAPTAIN", "FIRST_OFFICER"}:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=invalid_seat",
            status_code=HTTP_303_SEE_OTHER,
        )

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    new_user = db.query(User).filter(User.id == new_user_id).first()
    if new_user is None:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=user_not_found",
            status_code=HTTP_303_SEE_OTHER,
        )

    if new_user.role not in {"pilot", "copilot"}:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=invalid_crew_role",
            status_code=HTTP_303_SEE_OTHER,
        )

    opposite_seat = "FIRST_OFFICER" if seat == "CAPTAIN" else "CAPTAIN"
    opposite_active_assignment = (
        db.query(CrewAssignment)
        .filter(
            CrewAssignment.flight_id == flight_id,
            CrewAssignment.seat == opposite_seat,
            CrewAssignment.end_time.is_(None),
        )
        .first()
    )
    if opposite_active_assignment is not None and opposite_active_assignment.user_id == new_user_id:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?error=same_personnel",
            status_code=HTTP_303_SEE_OTHER,
        )

    active_assignment = (
        db.query(CrewAssignment)
        .filter(
            CrewAssignment.flight_id == flight_id,
            CrewAssignment.seat == seat,
            CrewAssignment.end_time.is_(None),
        )
        .first()
    )

    now = datetime.utcnow()

    if active_assignment and active_assignment.user_id == new_user_id:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/crew?success=crew_updated",
            status_code=HTTP_303_SEE_OTHER,
        )

    if active_assignment:
        active_assignment.end_time = now

    new_assignment = CrewAssignment(
        flight_id=flight_id,
        user_id=new_user_id,
        seat=seat,
        start_time=now,
    )
    db.add(new_assignment)
    db.commit()

    return RedirectResponse(
        url=f"/admin-ui/flights/{flight_id}/crew?success=crew_updated",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.get("/reports-ui", include_in_schema=False)
def reports_ui(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        total_flights = db.query(Flight).count()
        today_flights = db.query(Flight).filter(Flight.flight_date == date.today()).count()
        total_users = db.query(User).count()
        total_active_assignments = db.query(CrewAssignment).filter(CrewAssignment.end_time.is_(None)).count()

        admin_count = db.query(User).filter(User.role == "admin").count()
        pilot_count = db.query(User).filter(User.role == "pilot").count()
        copilot_count = db.query(User).filter(User.role == "copilot").count()
        technician_count = db.query(User).filter(User.role == "technician").count()

        recent_flights = db.query(Flight).order_by(Flight.id.desc()).limit(5).all()
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "reports.html",
        i18n_ctx(
            request,
            {
                "user": user,
                "total_flights": total_flights,
                "today_flights": today_flights,
                "total_users": total_users,
                "total_active_assignments": total_active_assignments,
                "admin_count": admin_count,
                "pilot_count": pilot_count,
                "copilot_count": copilot_count,
                "technician_count": technician_count,
                "recent_flights": recent_flights,
            },
        ),
    )


@app.get("/reports-ui/crew", include_in_schema=False)
def crew_reports_ui(
    request: Request,
    start_date: str = "",
    end_date: str = "",
    user_id: str = "",
    role: str = "",
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        filter_user_id = int(user_id) if user_id.strip().isdigit() else None
        filter_role = role.strip().lower() if role.strip() else None

        summary_rows, detail_rows = build_crew_report_data(
            db=db,
            start_date_str=start_date.strip() or None,
            end_date_str=end_date.strip() or None,
            user_id=filter_user_id,
            role=filter_role,
        )

        crew_users = (
            db.query(User)
            .filter(User.role.in_(["pilot", "copilot"]))
            .order_by(User.username)
            .all()
        )
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "crew_reports.html",
        i18n_ctx(
            request,
            {
                "user": current_user,
                "summary_rows": summary_rows,
                "detail_rows": detail_rows,
                "crew_users": crew_users,
                "filters": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "user_id": user_id,
                    "role": role,
                },
            },
        ),
    )


@app.get("/reports-ui/crew/export", include_in_schema=False)
def export_crew_reports(
    request: Request,
    start_date: str = "",
    end_date: str = "",
    user_id: str = "",
    role: str = "",
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        filter_user_id = int(user_id) if user_id.strip().isdigit() else None
        filter_role = role.strip().lower() if role.strip() else None

        summary_rows, detail_rows = build_crew_report_data(
            db=db,
            start_date_str=start_date.strip() or None,
            end_date_str=end_date.strip() or None,
            user_id=filter_user_id,
            role=filter_role,
        )
    finally:
        db.close()

    workbook = Workbook()

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    header_font = Font(bold=True)
    total_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    def style_sheet(ws):
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for column_cells in ws.columns:
            max_length = 0
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)

            for cell in column_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    if len(value) > max_length:
                        max_length = len(value)
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, 12), 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def style_total_row(ws, row_number: int):
        for cell in ws[row_number]:
            cell.fill = total_fill
            cell.font = total_font

    summary_total_flights = sum(row["total_flights"] for row in summary_rows)
    summary_total_assignments = sum(row["total_assignments"] for row in summary_rows)
    summary_total_seconds = sum(row["total_seconds"] for row in summary_rows)
    summary_total_hms = format_duration_hms(summary_total_seconds)

    detail_total_minutes = sum(row["duration_minutes"] for row in detail_rows)
    detail_total_seconds = sum(row["duration_seconds"] for row in detail_rows)
    detail_total_hms = format_duration_hms(detail_total_seconds)
    detail_total_segments = len(detail_rows)

    summary_sheet = workbook.active
    summary_sheet.title = "Crew Summary"
    summary_sheet.append(
        [
            "User ID",
            "Username",
            "Role",
            "Total Flights",
            "Total Assignments",
            "Total Minutes",
            "Total Time (HH:MM:SS)",
        ]
    )

    for row in summary_rows:
        summary_sheet.append(
            [
                row["user_id"],
                row["username"],
                row["role"],
                row["total_flights"],
                row["total_assignments"],
                row["total_minutes"],
                row["total_hms"],
            ]
        )

    summary_sheet.append(
        [
            "",
            "TOTAL",
            "",
            summary_total_flights,
            summary_total_assignments,
            summary_total_seconds // 60,
            summary_total_hms,
        ]
    )
    style_sheet(summary_sheet)
    style_total_row(summary_sheet, summary_sheet.max_row)

    detail_sheet = workbook.create_sheet(title="Assignment Details")
    detail_sheet.append(
        [
            "Flight ID",
            "Flight No",
            "Flight Date",
            "From",
            "To",
            "User ID",
            "Username",
            "Role",
            "Seat",
            "Start Time",
            "End Time",
            "Duration Minutes",
            "Duration (HH:MM:SS)",
            "Active",
        ]
    )

    for row in detail_rows:
        detail_sheet.append(
            [
                row["flight_id"],
                row["flight_no"],
                str(row["flight_date"]),
                row["departure_airport"],
                row["arrival_airport"],
                row["user_id"],
                row["username"],
                row["role"],
                row["seat"],
                str(row["start_time"]),
                str(row["end_time"]) if row["end_time"] else "",
                row["duration_minutes"],
                row["duration_hms"],
                "YES" if row["is_active"] else "NO",
            ]
        )

    detail_sheet.append(
        [
            "",
            "TOTAL",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            detail_total_minutes,
            detail_total_hms,
            f"{detail_total_segments} segments",
        ]
    )
    style_sheet(detail_sheet)
    style_total_row(detail_sheet, detail_sheet.max_row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename_parts = ["crew_reports"]
    if start_date.strip():
        filename_parts.append(start_date.strip())
    if end_date.strip():
        filename_parts.append(end_date.strip())
    if role.strip():
        filename_parts.append(role.strip())
    if user_id.strip():
        filename_parts.append(f"user_{user_id.strip()}")

    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/admin-ui/flights/{flight_id}/delete", include_in_schema=False)
def delete_flight(
    flight_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = get_session_user(request)
    if current_user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if current_user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    db.query(CrewAssignment).filter(CrewAssignment.flight_id == flight_id).delete()
    db.query(MaintenanceLog).filter(MaintenanceLog.flight_id == flight_id).delete()

    db.delete(flight)
    db.commit()

    return RedirectResponse(url="/admin-ui?success=flight_deleted", status_code=HTTP_303_SEE_OTHER)


@app.get("/admin-ui/flights/{flight_id}/edit", include_in_schema=False)
def edit_flight_page(flight_id: int, request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        flight = db.query(Flight).filter(Flight.id == flight_id).first()
        if flight is None:
            return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "edit_flight.html",
        i18n_ctx(request, {"user": user, "flight": flight}),
    )


@app.post("/admin-ui/flights/{flight_id}/edit", include_in_schema=False)
def edit_flight_from_form(
    flight_id: int,
    request: Request,
    flight_no: str = Form(...),
    flight_date: str = Form(...),
    departure_airport: str = Form(...),
    arrival_airport: str = Form(...),
    sched_dep: str = Form(...),
    sched_arr: str = Form(...),
    actual_dep: str = Form(""),
    actual_arr: str = Form(""),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role != "admin":
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/admin-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    parsed_sched_dep = datetime.fromisoformat(sched_dep)
    parsed_sched_arr = datetime.fromisoformat(sched_arr)
    parsed_actual_dep = datetime.fromisoformat(actual_dep) if actual_dep else None
    parsed_actual_arr = datetime.fromisoformat(actual_arr) if actual_arr else None

    if parsed_sched_arr < parsed_sched_dep:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/edit?error=invalid_schedule_range",
            status_code=HTTP_303_SEE_OTHER,
        )

    if parsed_actual_dep is not None and parsed_actual_dep < parsed_sched_dep:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/edit?error=invalid_actual_departure",
            status_code=HTTP_303_SEE_OTHER,
        )

    if parsed_actual_arr is not None and parsed_actual_arr < parsed_sched_arr:
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/edit?error=invalid_actual_arrival",
            status_code=HTTP_303_SEE_OTHER,
        )

    if (
        parsed_actual_dep is not None
        and parsed_actual_arr is not None
        and parsed_actual_arr < parsed_actual_dep
    ):
        return RedirectResponse(
            url=f"/admin-ui/flights/{flight_id}/edit?error=invalid_actual_range",
            status_code=HTTP_303_SEE_OTHER,
        )

    flight.flight_no = flight_no.strip()
    flight.flight_date = date.fromisoformat(flight_date)
    flight.departure_airport = departure_airport.strip().upper()
    flight.arrival_airport = arrival_airport.strip().upper()
    flight.sched_dep = parsed_sched_dep
    flight.sched_arr = parsed_sched_arr
    flight.actual_dep = parsed_actual_dep
    flight.actual_arr = parsed_actual_arr

    if parsed_actual_arr is not None:
        open_assignments = (
            db.query(CrewAssignment)
            .filter(
                CrewAssignment.flight_id == flight_id,
                CrewAssignment.end_time.is_(None),
            )
            .all()
        )

        for assignment in open_assignments:
            if parsed_actual_arr >= assignment.start_time:
                assignment.end_time = parsed_actual_arr

    db.commit()

    return RedirectResponse(url="/admin-ui?success=flight_updated", status_code=HTTP_303_SEE_OTHER)


@app.get("/flights-ui/{flight_id}/maintenance", include_in_schema=False)
def maintenance_page(flight_id: int, request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role not in {"admin", "technician"}:
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    db = SessionLocal()
    try:
        flight = db.query(Flight).filter(Flight.id == flight_id).first()
        if flight is None:
            return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

        logs = (
            db.query(MaintenanceLog)
            .options(joinedload(MaintenanceLog.user))
            .filter(MaintenanceLog.flight_id == flight_id)
            .order_by(MaintenanceLog.logged_at.desc())
            .all()
        )
    finally:
        db.close()

    return templates.TemplateResponse(
        request,
        "maintenance.html",
        i18n_ctx(
            request,
            {
                "user": user,
                "flight": flight,
                "logs": logs,
            },
        ),
    )


@app.post("/flights-ui/{flight_id}/maintenance", include_in_schema=False)
def create_maintenance_log(
    flight_id: int,
    request: Request,
    description: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    if user.role not in {"admin", "technician"}:
        return RedirectResponse(url="/flights-ui", status_code=HTTP_303_SEE_OTHER)

    flight = db.query(Flight).filter(Flight.id == flight_id).first()
    if flight is None:
        return RedirectResponse(url="/flights-ui?error=flight_not_found", status_code=HTTP_303_SEE_OTHER)

    description = description.strip()
    if not description:
        return RedirectResponse(
            url=f"/flights-ui/{flight_id}/maintenance?error=empty_description",
            status_code=HTTP_303_SEE_OTHER,
        )

    log = MaintenanceLog(
        flight_id=flight_id,
        user_id=user.id,
        description=description,
    )
    db.add(log)
    db.commit()

    return RedirectResponse(
        url=f"/flights-ui/{flight_id}/maintenance?success=maintenance_created",
        status_code=HTTP_303_SEE_OTHER,
    )


@app.get("/change-password")
def change_password_page(request: Request):
    user = get_session_user(request)
    if user is None:
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse(
        request,
        "change_password.html",
        i18n_ctx(request, {"user": user})
    )

@app.post("/change-password")
def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_session_user(request)
    lang = get_locale(request)
    t = lambda k: TRANSLATIONS[lang].get(k, k)

    if user is None:
        return RedirectResponse("/login", status_code=303)

    if not pwd_context.verify(current_password, user.password_hash):
        return templates.TemplateResponse(
            request,
            "change_password.html",
            i18n_ctx(request, {"user": user, "error": t("password.error.wrong")})
        )
    if new_password != confirm_password:
        return templates.TemplateResponse(
            request,
            "change_password.html",
            i18n_ctx(request, {"user": user, "error": t("password.error.mismatch")})
        )
    user.password_hash = pwd_context.hash(new_password)
    db.add(user)
    db.commit()
    return templates.TemplateResponse(
        request,
        "change_password.html",
        i18n_ctx(request, {"user": user, "success": t("password.success")})
    )





@app.get("/health")
def health():
    return {"status": "ok"}
